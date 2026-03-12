#!/usr/bin/env python3
"""
EZproxy Log File Bulk Analyzer
Parses and analyzes EZproxy access logs to produce comprehensive summary statistics,
top users, top resources, traffic trends, and error reports. Supports multiple input
log formats and exports results to text, CSV, JSON, or Excel formats.

Input log formats supported:
  - EZproxy default: "%h %l %u %t \"%r\" %s %b"
  - EZproxy extended (with referrer and user-agent)
  - SPU (Starting Point URL) log format

Output formats:
  - text: Formatted text report (default, to stdout or file)
  - csv:  Comma-separated values with sections for each data category
  - json: Structured JSON with all metrics and trending data
  - xlsx: Multi-sheet Excel workbook with formatted tables

Usage:
  python ezproxy_analyzer.py [options] <logfile_or_directory> [logfile2 ...]

Examples:
  python ezproxy_analyzer.py /var/log/ezproxy/
  python ezproxy_analyzer.py ezproxy.log.2024-01 ezproxy.log.2024-02
  python ezproxy_analyzer.py --output report.csv --format csv /logs/
  python ezproxy_analyzer.py --output report.xlsx --format xlsx --top 20 /logs/
  python ezproxy_analyzer.py --output report.json --format json /logs/
  python ezproxy_analyzer.py --top 20 --since 2024-01-01 --until 2024-12-31 /logs/
  python ezproxy_analyzer.py --exclude-users exclude_list.txt --format xlsx /logs/
"""

import re
import os
import sys
import csv
import gzip
import argparse
import json
from pathlib import Path
from datetime import datetime, timedelta
from collections import defaultdict, Counter
from typing import Iterator

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Log line regex patterns
# ---------------------------------------------------------------------------

# Standard EZproxy / Combined Log Format
# 1.2.3.4 - username [01/Jan/2024:00:00:00 -0500] "GET http://... HTTP/1.1" 200 1234
#
# The request field uses  (?:[^"\\]|\\.)*  so it matches:
#   - any char that is not a quote or backslash  [^"\\]
#   - OR a backslash followed by any char        \\.
# This handles lines where the request contains escaped quotes like
#   "{\"id\": ...}"  (JSON-RPC probes, malformed bots, etc.)
COMBINED_RE = re.compile(
    r'(?P<host>\S+)\s+'
    r'(?P<ident>\S+)\s+'
    r'(?P<user>\S+)\s+'
    r'\[(?P<time>[^\]]+)\]\s+'
    r'"(?P<request>(?:[^"\\]|\\.)*?)"\s+'
    r'(?P<status>\d{3})\s+'
    r'(?P<bytes>\S+)'
    r'(?:\s+"(?P<referer>(?:[^"\\]|\\.)*)")?'
    r'(?:\s+"(?P<agent>(?:[^"\\]|\\.)*)")?'
)

# Looser fallback: same prefix but status/bytes are optional.
# Catches malformed lines that have a valid IP + timestamp but a
# truncated or otherwise broken tail (e.g. very long URLs cut off).
COMBINED_LOOSE_RE = re.compile(
    r'(?P<host>\S+)\s+'
    r'(?P<ident>\S+)\s+'
    r'(?P<user>\S+)\s+'
    r'\[(?P<time>[^\]]+)\]\s+'
    r'"(?P<request>(?:[^"\\]|\\.)*?)"'
    r'(?:\s+(?P<status>\d{3}))?'
    r'(?:\s+(?P<bytes>\S+))?'
)

# SPU (Starting Point URL) format:  username timestamp url session
SPU_RE = re.compile(
    r'^(?P<user>\S+)\s+'
    r'(?P<time>\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\s+'
    r'(?P<url>\S+)\s*'
    r'(?P<session>\S*)'
)

TIMESTAMP_FORMATS = [
    "%d/%b/%Y:%H:%M:%S %z",   # CLF: 01/Jan/2024:12:00:00 -0500
    "%d/%b/%Y:%H:%M:%S",       # CLF without tz
    "%Y-%m-%d %H:%M:%S",       # ISO-like (SPU)
]


def parse_timestamp(raw: str) -> datetime | None:
    for fmt in TIMESTAMP_FORMATS:
        try:
            return datetime.strptime(raw.strip(), fmt)
        except ValueError:
            continue
    return None


def extract_host(url: str) -> str:
    """Pull the database/resource hostname from a proxied URL."""
    m = re.search(r'https?://([^/:?\s]+)', url)
    return m.group(1).lower() if m else url.split()[1] if ' ' in url else url


# ---------------------------------------------------------------------------
# Log parsers
# ---------------------------------------------------------------------------

def _build_combined_record(m) -> dict:
    """Shared post-processing for both strict and loose combined matches."""
    d = m.groupdict()
    d['timestamp'] = parse_timestamp(d['time'])
    # Unescape backslash sequences so lines like  "{\"id\": ...}"
    # are stored as clean text for URL extraction.
    raw_req = d.get('request') or ''
    try:
        d['request'] = raw_req.encode('raw_unicode_escape').decode('unicode_escape')
    except Exception:
        d['request'] = raw_req
    d['resource'] = extract_host(d['request'])
    d['bytes_int'] = int(d['bytes']) if d.get('bytes') and d['bytes'] != '-' else 0
    d.setdefault('status', '-')
    return d


def parse_line(line: str) -> dict | None:
    """Try strict CLF, then loose CLF fallback, then SPU."""
    line = line.rstrip('\n')
    if not line or line.startswith('#'):
        return None

    # 1. Strict CLF / combined (requires status + bytes)
    m = COMBINED_RE.match(line)
    if m:
        return _build_combined_record(m)

    # 2. Loose CLF fallback: handles truncated lines, missing status/bytes,
    #    and probes whose request field contains escaped quotes or JSON.
    m = COMBINED_LOOSE_RE.match(line)
    if m:
        return _build_combined_record(m)

    # 3. Last-resort CLF: line was truncated before the closing quote.
    #    Match IP + timestamp + whatever is inside an unclosed quote field.
    m = re.match(
        r'(?P<host>\S+)\s+(?P<ident>\S+)\s+(?P<user>\S+)\s+'
        r'\[(?P<time>[^\]]+)\]\s+"(?P<request>.+)$',
        line
    )
    if m:
        d = m.groupdict()
        d['status'] = '-'
        d['bytes'] = '-'
        d['timestamp'] = parse_timestamp(d['time'])
        d['request'] = d.get('request', '')
        d['resource'] = extract_host(d['request'])
        d['bytes_int'] = 0
        return d

    # 4. SPU format
    m = SPU_RE.match(line)
    if m:
        d = m.groupdict()
        d['host'] = '-'
        d['status'] = '200'
        d['bytes_int'] = 0
        d['request'] = d.get('url', '')
        d['timestamp'] = parse_timestamp(d['time'])
        d['resource'] = extract_host(d.get('url') or '')
        return d

    return None


def open_log(path: Path):
    """Open plain or gzip-compressed log files."""
    if path.suffix in ('.gz', '.gzip'):
        return gzip.open(path, 'rt', errors='replace')
    return open(path, 'r', errors='replace')


def iter_log_files(paths: list[Path]) -> Iterator[Path]:
    """Yield individual log files from a mix of files and directories."""
    for p in paths:
        if p.is_dir():
            for f in sorted(p.iterdir()):
                if f.is_file() and not f.name.startswith('.'):
                    yield f
        elif p.is_file():
            yield p
        else:
            print(f"[WARNING] Path not found: {p}", file=sys.stderr)


# ---------------------------------------------------------------------------
# Analyzer
# ---------------------------------------------------------------------------

class EZproxyAnalyzer:
    def __init__(self, since: datetime | None = None, until: datetime | None = None,
                 exclude_users: set | None = None, exclude_hosts: set | None = None):
        self.since = since
        self.until = until
        self.exclude_users = exclude_users or set()
        self.exclude_hosts = exclude_hosts or set()

        # Counters
        self.total_lines = 0
        self.parsed_lines = 0
        self.skipped_lines = 0

        self.user_counter: Counter = Counter()
        self.resource_counter: Counter = Counter()
        self.status_counter: Counter = Counter()
        self.hourly_counter: Counter = Counter()   # key: "YYYY-MM-DD HH"
        self.daily_counter: Counter = Counter()    # key: "YYYY-MM-DD"
        self.monthly_counter: Counter = Counter()  # key: "YYYY-MM"
        self.ip_counter: Counter = Counter()
        self.bytes_by_resource: Counter = Counter()
        self.bytes_total: int = 0
        self.error_lines: list[str] = []

        self.first_ts: datetime | None = None
        self.last_ts: datetime | None = None
        self.files_processed: list[str] = []

    def ingest(self, paths: list[Path]) -> None:
        for log_path in iter_log_files(paths):
            self.files_processed.append(str(log_path))
            print(f"  Processing: {log_path}", file=sys.stderr)
            try:
                with open_log(log_path) as fh:
                    for raw_line in fh:
                        self._process_line(raw_line)
            except Exception as e:
                print(f"  [ERROR] Could not read {log_path}: {e}", file=sys.stderr)

    def _process_line(self, raw: str) -> None:
        self.total_lines += 1
        record = parse_line(raw)

        if record is None:
            self.skipped_lines += 1
            if len(self.error_lines) < 20:
                self.error_lines.append(raw.strip())
            return

        ts = record.get('timestamp')
        if ts:
            if self.since and ts < self.since:
                return
            if self.until and ts > self.until:
                return
            if self.first_ts is None or ts < self.first_ts:
                self.first_ts = ts
            if self.last_ts is None or ts > self.last_ts:
                self.last_ts = ts

        user = record.get('user', '-')
        host = record.get('host', '-')
        resource = record.get('resource', '-')
        status = record.get('status', '-')
        byt = record.get('bytes_int', 0)

        if user in self.exclude_users or host in self.exclude_hosts:
            return

        self.parsed_lines += 1
        self.user_counter[user] += 1
        self.resource_counter[resource] += 1
        self.status_counter[status] += 1
        self.ip_counter[host] += 1
        self.bytes_by_resource[resource] += byt
        self.bytes_total += byt

        if ts:
            self.hourly_counter[ts.strftime('%Y-%m-%d %H')] += 1
            self.daily_counter[ts.strftime('%Y-%m-%d')] += 1
            self.monthly_counter[ts.strftime('%Y-%m')] += 1

    # -----------------------------------------------------------------------
    # Reporting helpers
    # -----------------------------------------------------------------------

    def _fmt_bytes(self, b: int | float) -> str:
        for unit in ('B', 'KB', 'MB', 'GB', 'TB'):
            if b < 1024:
                return f"{b:.1f} {unit}"
            b /= 1024
        return f"{b:.1f} PB"

    def summary(self) -> dict:
        return {
            'files_processed': len(self.files_processed),
            'total_lines': self.total_lines,
            'parsed_lines': self.parsed_lines,
            'skipped_lines': self.skipped_lines,
            'unique_users': len(self.user_counter),
            'unique_resources': len(self.resource_counter),
            'unique_ips': len(self.ip_counter),
            'total_bytes': self.bytes_total,
            'total_bytes_human': self._fmt_bytes(self.bytes_total),
            'date_range_start': str(self.first_ts) if self.first_ts else 'N/A',
            'date_range_end': str(self.last_ts) if self.last_ts else 'N/A',
        }

    # -----------------------------------------------------------------------
    # Output formatters
    # -----------------------------------------------------------------------

    def print_report(self, top_n: int | None = None) -> None:
        s = self.summary()
        width = 60
        top_label = str(top_n) if top_n is not None else 'All'

        print("=" * width)
        print("  EZproxy Log Analysis Report")
        print("=" * width)
        print(f"  Files processed  : {s['files_processed']}")
        print(f"  Date range       : {s['date_range_start']}  →  {s['date_range_end']}")
        print(f"  Total log lines  : {s['total_lines']:,}")
        print(f"  Parsed lines     : {s['parsed_lines']:,}")
        print(f"  Skipped / bad    : {s['skipped_lines']:,}")
        print(f"  Unique users     : {s['unique_users']:,}")
        print(f"  Unique resources : {s['unique_resources']:,}")
        print(f"  Unique IPs       : {s['unique_ips']:,}")
        print(f"  Total transfer   : {s['total_bytes_human']}")
        print()

        self._print_section(f"{top_label} Users by Requests",
                            self.user_counter, top_n)

        self._print_section(f"{top_label} Resources (Hostnames) by Requests",
                            self.resource_counter, top_n,
                            extra=self.bytes_by_resource,
                            extra_label="Transfer")

        self._print_section("HTTP Status Codes", self.status_counter, None)

        self._print_section(f"{top_label} Client IPs",
                            self.ip_counter, top_n)

        self._print_trend("Monthly Traffic", self.monthly_counter)
        self._print_trend("Daily Traffic (last 30 days)",
                          self.daily_counter, limit=30)
        self._print_trend("Hourly Distribution (requests per hour of day)",
                          self._collapse_hours())

        if self.error_lines:
            print("-" * width)
            print(f"  Sample unparseable lines (first {len(self.error_lines)}):")
            for ln in self.error_lines[:5]:
                print(f"    {ln[:120]}")
        print("=" * width)

    def _print_section(self, title, counter, n, extra=None, extra_label=''):
        print(f"--- {title} " + "-" * max(1, 56 - len(title)))
        # most_common(None) returns all items sorted by count
        for item, count in counter.most_common(n):
            bar = '#' * min(30, count * 30 // max(counter.values(), default=1))
            suffix = ''
            if extra and item in extra:
                suffix = f"  [{self._fmt_bytes(extra[item])} {extra_label}]"
            print(f"  {str(item):<40} {count:>8,}  {bar}{suffix}")
        print()

    def _print_trend(self, title, counter, limit=None):
        print(f"--- {title} " + "-" * max(1, 56 - len(title)))
        items = sorted(counter.items())
        if limit:
            items = items[-limit:]
        max_val = max((v for _, v in items), default=1)
        for key, val in items:
            bar = '#' * min(40, val * 40 // max_val)
            print(f"  {key}  {val:>8,}  {bar}")
        print()

    def _collapse_hours(self) -> Counter:
        """Sum requests by hour-of-day (0–23) across all dates."""
        hod: Counter = Counter()
        for key, val in self.hourly_counter.items():
            hour = key.split()[-1]
            hod[hour] += val
        return hod

    # -----------------------------------------------------------------------
    # CSV export
    # -----------------------------------------------------------------------

    def write_csv(self, output_path: str, top_n: int | None = None) -> None:
        with open(output_path, 'w', newline='', encoding='utf-8') as fh:
            writer = csv.writer(fh)
            top_label = str(top_n) if top_n is not None else 'All'

            # Summary
            writer.writerow(['=== SUMMARY ==='])
            for k, v in self.summary().items():
                writer.writerow([k, v])
            writer.writerow([])

            # All users
            writer.writerow([f'=== {top_label} USERS ==='])
            writer.writerow(['Username', 'Requests'])
            for user, cnt in self.user_counter.most_common(top_n):
                writer.writerow([user, cnt])
            writer.writerow([])

            # All resources
            writer.writerow([f'=== {top_label} RESOURCES ==='])
            writer.writerow(['Resource', 'Requests', 'Bytes'])
            for res, cnt in self.resource_counter.most_common(top_n):
                writer.writerow([res, cnt, self.bytes_by_resource.get(res, 0)])
            writer.writerow([])

            # Status codes
            writer.writerow(['=== STATUS CODES ==='])
            writer.writerow(['Status', 'Count'])
            for status, cnt in sorted(self.status_counter.items()):
                writer.writerow([status, cnt])
            writer.writerow([])

            # Daily trend
            writer.writerow(['=== DAILY TRAFFIC ==='])
            writer.writerow(['Date', 'Requests'])
            for day, cnt in sorted(self.daily_counter.items()):
                writer.writerow([day, cnt])

        print(f"CSV report written to: {output_path}", file=sys.stderr)

    # -----------------------------------------------------------------------
    # JSON export
    # -----------------------------------------------------------------------

    def write_json(self, output_path: str, top_n: int | None = None) -> None:
        data = {
            'summary': self.summary(),
            'users': self.user_counter.most_common(top_n),
            'resources': [
                {'resource': r, 'requests': c,
                 'bytes': self.bytes_by_resource.get(r, 0)}
                for r, c in self.resource_counter.most_common(top_n)
            ],
            'status_codes': dict(self.status_counter),
            'ips': self.ip_counter.most_common(top_n),
            'monthly_traffic': dict(sorted(self.monthly_counter.items())),
            'daily_traffic': dict(sorted(self.daily_counter.items())),
        }
        with open(output_path, 'w', encoding='utf-8') as fh:
            json.dump(data, fh, indent=2)
        print(f"JSON report written to: {output_path}", file=sys.stderr)

    # -----------------------------------------------------------------------
    # Excel export
    # -----------------------------------------------------------------------

    def write_excel(self, output_path: str, top_n: int | None = None) -> None:
        if not HAS_OPENPYXL:
            print("[ERROR] openpyxl is not installed. Install it with: pip install openpyxl", 
                  file=sys.stderr)
            return

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        
        # Summary sheet
        self._excel_summary(wb, header_fill, header_font)
        
        # Users sheet
        self._excel_users(wb, top_n, header_fill, header_font)
        
        # Resources sheet
        self._excel_resources(wb, top_n, header_fill, header_font)
        
        # Status codes sheet
        self._excel_status_codes(wb, header_fill, header_font)
        
        # Monthly traffic sheet
        self._excel_monthly(wb, header_fill, header_font)
        
        # Daily traffic sheet
        self._excel_daily(wb, header_fill, header_font)
        
        # IPs sheet
        self._excel_ips(wb, top_n, header_fill, header_font)
        
        wb.save(output_path)
        print(f"Excel report written to: {output_path}", file=sys.stderr)

    def _excel_summary(self, wb, header_fill, header_font):
        ws = wb.create_sheet("Summary", 0)
        s = self.summary()
        ws.append(["Metric", "Value"])
        self._excel_format_headers(ws, header_fill, header_font)
        for k, v in s.items():
            ws.append([k, v])
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50

    def _excel_users(self, wb, top_n, header_fill, header_font):
        ws = wb.create_sheet("Users")
        ws.append(["User", "Requests"])
        self._excel_format_headers(ws, header_fill, header_font)
        for user, count in self.user_counter.most_common(top_n):
            ws.append([user, count])
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15

    def _excel_resources(self, wb, top_n, header_fill, header_font):
        ws = wb.create_sheet("Resources")
        ws.append(["Resource", "Requests", "Bytes"])
        self._excel_format_headers(ws, header_fill, header_font)
        for res, count in self.resource_counter.most_common(top_n):
            bytes_val = self.bytes_by_resource.get(res, 0)
            ws.append([res, count, bytes_val])
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    def _excel_status_codes(self, wb, header_fill, header_font):
        ws = wb.create_sheet("Status Codes")
        ws.append(["Status", "Count"])
        self._excel_format_headers(ws, header_fill, header_font)
        for status, count in sorted(self.status_counter.items()):
            ws.append([status, count])
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15

    def _excel_monthly(self, wb, header_fill, header_font):
        ws = wb.create_sheet("Monthly Traffic")
        ws.append(["Month", "Requests"])
        self._excel_format_headers(ws, header_fill, header_font)
        for month, count in sorted(self.monthly_counter.items()):
            ws.append([month, count])
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15

    def _excel_daily(self, wb, header_fill, header_font):
        ws = wb.create_sheet("Daily Traffic")
        ws.append(["Date", "Requests"])
        self._excel_format_headers(ws, header_fill, header_font)
        for day, count in sorted(self.daily_counter.items()):
            ws.append([day, count])
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15

    def _excel_ips(self, wb, top_n, header_fill, header_font):
        ws = wb.create_sheet("Client IPs")
        ws.append(["IP Address", "Requests"])
        self._excel_format_headers(ws, header_fill, header_font)
        for ip, count in self.ip_counter.most_common(top_n):
            ws.append([ip, count])
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

    def _excel_format_headers(self, ws, fill, font):
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(
        description='Analyze EZproxy log files in bulk.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    parser.add_argument('paths', nargs='+', metavar='PATH',
                        help='Log file(s) or directory/directories to analyze')
    parser.add_argument('--top', type=int, default=None, metavar='N',
                        help='Limit output to top N entries per category (default: all)')
    parser.add_argument('--since', metavar='YYYY-MM-DD',
                        help='Only include log entries on or after this date')
    parser.add_argument('--until', metavar='YYYY-MM-DD',
                        help='Only include log entries on or before this date')
    parser.add_argument('--exclude-users', metavar='FILE',
                        help='Path to a text file with one username per line to exclude')
    parser.add_argument('--exclude-hosts', metavar='FILE',
                        help='Path to a text file with one IP/hostname per line to exclude')
    parser.add_argument('--output', '-o', metavar='FILE',
                        help='Write report to file (.csv or .json); stdout if omitted')
    parser.add_argument('--format', choices=['text', 'csv', 'json', 'xlsx'], default='text',
                        help='Output format (default: text)')
    return parser.parse_args()


def load_exclusion_list(path: str | None) -> set:
    if not path:
        return set()
    try:
        with open(path) as fh:
            return {line.strip() for line in fh if line.strip()}
    except FileNotFoundError:
        print(f"[WARNING] Exclusion file not found: {path}", file=sys.stderr)
        return set()


def main():
    args = parse_args()

    since = datetime.strptime(args.since, '%Y-%m-%d') if args.since else None
    until = (datetime.strptime(args.until, '%Y-%m-%d') + timedelta(days=1)
             if args.until else None)

    analyzer = EZproxyAnalyzer(
        since=since,
        until=until,
        exclude_users=load_exclusion_list(args.exclude_users),
        exclude_hosts=load_exclusion_list(args.exclude_hosts),
    )

    paths = [Path(p) for p in args.paths]
    print("Scanning log files...", file=sys.stderr)
    analyzer.ingest(paths)
    print(f"Done. {analyzer.parsed_lines:,} records processed.\n", file=sys.stderr)

    fmt = args.format
    out = args.output

    if fmt == 'csv':
        if not out:
            out = 'ezproxy_report.csv'
        analyzer.write_csv(out, top_n=args.top)
    elif fmt == 'json':
        if not out:
            out = 'ezproxy_report.json'
        analyzer.write_json(out, top_n=args.top)
    elif fmt == 'xlsx':
        if not out:
            out = 'ezproxy_report.xlsx'
        analyzer.write_excel(out, top_n=args.top)
    else:
        if out:
            sys.stdout = open(out, 'w', encoding='utf-8')
        analyzer.print_report(top_n=args.top)


if __name__ == '__main__':
    main()
